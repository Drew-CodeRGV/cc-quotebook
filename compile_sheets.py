#!/usr/bin/env python
#
#
# Requested by Drew-CodeRGV
# Created by ldartez
# 
#
#
# compile master sheet with required information
# master sheet columns: 
#   * Manufacturer
#   * Type
#   * Part Number
#   * Description
#   * List Price
#
# Aruba Mapping:
#   Manufacturer: Aruba
#   Type: Sheet Name (Access points, Switches, Central Licensing)
#   Part Number: Column A value
#   Description: Column B value
#   List Price: Column C value
#
# Cradlepoint Mapping:
#   Manufacturer: Cradlepoint
#   Type: Column B value
#   Part Number: Column D value
#   Description: Column G value
#   List Price: Column F value
#
# Fortinet Mapping:
#   Manufacturer: Fortinet
#   Type: Sheet Name (FortiGate, Wireless Products)
#   Part Number: Column B value
#   Description: Column C value
#   List Price: Column E value
#
# Meraki Mapping:
#   Manufacturer: Meraki
#   Type: Column B value
#   Part Number: Column C value
#   Description: Column D value
#   List Price: Column F value
#
# SnapAV Mapping: (Note: use filter pre-selected for POWER)
#   Manufacturer: SnapAV
#   Type: Column A value
#   Part Number: Column B value
#   Description: Column C value
#   List Price: Column J value

def filter_cells(filt, x):
    '''Return true if x does not include any substrings in filt
    '''
    passing = True
    for val in x:
        for fv in filt:
            if str(fv) in str(val):
                passing = False
                break
        if not passing:
            break
    return passing

def parse_workbook_aruba(wbin):
    '''
    Aruba Mapping:
    Manufacturer: Aruba
    Type: Sheet Name (Access points, Switches, Central Licensing)
    Part Number: Column A value
    Description: Column B value
    List Price: Column C value

    Row Filters: 'Indoor Access Points', 'Mounting Brackets',
                 'Outdoor Access Points', None
    '''
    sheets = ['Access Points', 'Switches', 'Central Licensing']
    row_filters = ['Indoor Access Points', 'Mounting Brackets',
            'Outdoor Access Points', 'Part Number', 'Series', 'None' ]
    mftr = 'Aruba'
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            for row in sheet.values:
                filtpass = filter_cells(row_filters, list(row))
                if not filtpass:
                    continue
                else:
                    row_out = (mftr, sheet.title, row[0], row[1], row[2])
                    result.append(row_out)
    return result

def parse_workbook_cradlepoint(wbin):
    '''
    Cradlepoint Mapping:
    Manufacturer: Cradlepoint
    Type: Column B value
    Part Number: Column D value
    Description: Column G value
    List Price: Column F value
    '''
    sheets = ['USA']
    row_filters = ['Cradlepoint USA MSRP', 'Company Confidential',
            'Products']
    types = ['Routers', 'Access Points', 'LTE Adapters', 'Performance Routers',
            'Virtual Router', 'Mobile First Responder Packages',
            'Gateways', 'FIPS', 'NetCloud', 'Threat Management',
            'Internet Security', 'Feature Licenses', 'Modems',
            'SIM-in-Box', 'Antennas', 'Cradlepoint Certified',
            'Power Supplies', 'Miscellaneous', 'COR Series Routers',
            'Accessories', 'AER Series Routers', 'Home Office',
            'M2M']
    mftr = 'Cradlepoint'
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            for row in sheet.values:
                if str(row[3]) in ['None', 'Note', 'Part Number']:
                    continue
                filtpass = filter_cells(types, str(row[1]).split())
                if not filtpass:
                    cur_type = row[1]
                row_out = (mftr, cur_type, row[3], row[6], row[5])
                result.append(row_out)
    return result

def parse_workbook_fortinet(wbin):
    '''
    Fortinet Mapping:
    Manufacturer: Fortinet
    Type: Sheet Name (FortiGate, Wireless Products)
    Part Number: Column B value
    Description: Column C value
    List Price: Column E value

    '''
    sheets = ['FortiGate', 'Wireless Products']
    row_filters = ['None', 'SKU', 'PRMA', 'Requires','HYPERLINK']
    mftr = 'Fortinet'
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            for row in sheet.values:
                filtpass = filter_cells(row_filters, str(row[1]).split())
                if not filtpass:
                    continue
                row_out = (mftr, s, row[1], row[2], row[4])
                result.append(row_out)
    return result
    
def parse_workbook_meraki(wbin):
    '''

    Meraki Mapping:
    Manufacturer: Meraki
    Type: Column B value
    Part Number: Column C value
    Description: Column D value
    List Price: Column F value
    '''
    sheets = ['Report']
    row_filters = ['Cisco']
    mftr = 'Meraki'
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            # skip first few rows
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if not filter_cells(row_filters, [row[1]]):
                    cur_type = str(row[1])
                    continue
                row_out = (mftr, cur_type, row[2], row[3], row[5])
                result.append(row_out)
    return result

def parse_workbook_snapav(wbin):
    '''
    SnapAV Mapping: (Note: use filter pre-selected for POWER)
    Manufacturer: SnapAV
    Type: Column A value
    Part Number: Column B value
    Description: Column C value
    List Price: Column J value

    '''
    sheets = ['Sheet 1']
    row_filters = ['Cisco']
    mftr = 'SnapAV'
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            # skip first few rows
            for row in sheet.iter_rows(values_only=True):
                if not 'Power' == str(row[0]):
                    continue
                row_out = (mftr, row[0], row[1], row[2], row[9])
                result.append(row_out)
    return result

   
if __name__ == "__main__":
    import argparse
    from os.path import abspath, basename, isdir, isfile, exists, join
    from os import walk, getcwd
    from openpyxl import load_workbook, Workbook
    from glob import glob

    p = argparse.ArgumentParser()
    p.add_argument('-o', '--output', default='master.xlsx',
            help='path to output file')
    p.add_argument('infiles', nargs='+')
    args = p.parse_args()
    
    fnames = []
    for fin in args.infiles:
        if not exists(fin):
            raise ValueError("file not found: {}".format(fin))
        elif isdir(fin):
            flist = glob(join(fin,'*.xlsx'))
            for f in flist:
                if not f in fnames:
                    fnames.append(f)
        elif isfile(fin):
            if not fin in fnames:
                fnames.append(fin)
    dout = []
    for f in fnames:
        if 'aruba' in basename(f).lower():
            msg = "Processing Aruba file: {}".format(f)
            print(msg)
            wb = load_workbook(f)
            rows = parse_workbook_aruba(wb)
            dout.extend(rows)
            wb.close()
        elif 'cradlepoint' in basename(f).lower():
            msg = "Processing Cradlepoint file: {}".format(f)
            print(msg)
            wb = load_workbook(f)
            rows = parse_workbook_cradlepoint(wb)
            dout.extend(rows)
            wb.close()
        elif 'fortinet' in basename(f).lower():
            msg = "Processing Fortinet file: {}".format(f)
            print(msg)
            wb = load_workbook(f)
            rows = parse_workbook_fortinet(wb)
            dout.extend(rows)
            wb.close()
        elif 'meraki' in basename(f).lower():
            msg = "Processing Meraki file: {}".format(f)
            print(msg)
            wb = load_workbook(f)
            rows = parse_workbook_meraki(wb)
            dout.extend(rows)
            wb.close()
        elif 'snapav' in basename(f).lower():
            msg = "Processing SnapAV file: {}".format(f)
            print(msg)
            wb = load_workbook(f)
            rows = parse_workbook_snapav(wb)
            dout.extend(rows)
            wb.close()

    wbout = Workbook()
    wsout = wbout.active
    
    hdr = ('Manufacturer', 'Type', 'Part Number', 'Description', 'List Price')
    wsout.append(hdr)
    if dout:
        for r in dout:
            wsout.append(r)
        wbout.save(args.output)
    wbout.close()



